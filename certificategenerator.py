import cv2 as cv
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.mime.image import MIMEImage


mail_content = '''Dear Participant,

" Type you want in Header "

'''


template_path = '/Your_Path/certificate_generator/"Edit Your Template".png'
details_path = '/Your_Path/certificate_generator/details.xlsx'
output_path = '/Your_Path/certificate_generator/certificates/'


sender_address = 'Your Mail Id@gmail.com'
sender_pass = 'Your Mail Id Password'


font_size = 1.25
font_color = (20, 20, 20)


coordinate_y_adjustment = 110
coordinate_x_adjustment = 90


obj = openpyxl.load_workbook(details_path)
sheet = obj.active


for i in range(2,4):

    get_name = sheet.cell(row = i ,column = 1)
    certi_name = get_name.value
    get_email = sheet.cell(row = i ,column = 2)
    receiver_address = get_email.value
    
    
    img = cv.imread(template_path)

    font = cv.FONT_HERSHEY_COMPLEX

    text_size = cv.getTextSize(certi_name, font, font_size, 10)[0]
    text_x = (img.shape[1] - text_size[0]) / 2 + coordinate_x_adjustment
    text_y = (img.shape[0] + text_size[1]) / 2 - coordinate_y_adjustment
    text_x = int(text_x)
    text_y = int(text_y)



    cv.putText(img, certi_name, (text_x ,text_y ), font, font_size, font_color, 2)

    certi_path = output_path + certi_name + '.png'

    cv.imwrite(certi_path,img)
    
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = receiver_address
    message['Subject'] = '" Type you want in Subject "'
    message.attach(MIMEText(mail_content, 'plain'))
    attach_file_name = output_path + certi_name + '.png'
    name = certi_name + '.png'
    print(attach_file_name)
    attach_file = open(attach_file_name, 'rb')
    png = MIMEImage ((attach_file).read())
    encoders.encode_base64(png)
    png.add_header('Content-Disposition', 'attachment' , filename=name)
    message.attach(png)
    
    session = smtplib.SMTP('smtp.gmail.com', 587)
    session.starttls()
    session.login(sender_address, sender_pass)
    session.sendmail(sender_address, receiver_address, message.as_string())
    session.quit()
    print('Mail Sent')

cv.destroyAllWindows()
